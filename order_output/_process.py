
import openpyxl, re, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def T(s='thin'): return Side(style=s)
def B(t=None,b=None,l=None,r=None): return Border(top=t,bottom=b,left=l,right=r)

def clean_names(raw):
    result = []
    for p in str(raw or '').split('▶'):
        if not p.strip(): continue
        q = re.search(r'\((\d+)개\)', p)
        qty = int(q.group(1)) if q else 1
        m = re.match(r'\s*(.+?)\s*,\s*[^,]*\(\d+개\)', p)
        name = m.group(1).strip() if m else (p[:p.find(',')].strip() if ',' in p else p.strip())
        if name and len(name) > 1:
            result.append((name, qty))
    return result

def is_eco(a): return '에코델타' in str(a) or '에코대로' in str(a)

wb_in = openpyxl.load_workbook(sys.argv[1], data_only=True)
ws_in = wb_in.active
hdr = [str(v or '').strip() for v in next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))]
rows = [{hdr[j]: r[j] for j in range(min(len(hdr),len(r)))} for r in ws_in.iter_rows(min_row=2, values_only=True) if any(r)]
print(f"데이터 {len(rows)}행 로드")

today_str = datetime.now().strftime('%Y년 %m월 %d일 (%a)')

# ── 소분작업 ──────────────────────────────────────────────────────
pm = {}
for row in rows:
    raw = str(row.get('상품명(타입제거)') or row.get('상품명') or '')
    base = int(row.get('주문수량') or row.get('수량') or 1)
    items = clean_names(raw) or [('(기타)', base)]
    for name, qty in items:
        pm[name] = pm.get(name, 0) + qty
sp = sorted(pm.items(), key=lambda x: -x[1])

wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = '소분작업'
ws1.merge_cells('A1:C1')
c=ws1['A1']; c.value=f'소분 작업 목록  ─  {today_str}'
c.font=Font(name='맑은 고딕',bold=True,size=14,color='FFFFFF')
c.fill=PatternFill('solid',fgColor='2E7D32')
c.alignment=Alignment(horizontal='center',vertical='center')
ws1.row_dimensions[1].height=32
ws1.append([])
ws1.append(['번호','상품명','수량'])
ws1.row_dimensions[3].height=24
for col,w in [(1,8),(2,48),(3,10)]: ws1.column_dimensions[chr(64+col)].width=w
for cell in ws1[3]:
    cell.font=Font(name='맑은 고딕',bold=True,size=12)
    cell.fill=PatternFill('solid',fgColor='C8E6C9')
    cell.alignment=Alignment(horizontal='center',vertical='center')
    cell.border=B(T('medium'),T('medium'),T('medium'),T('medium'))
for i,(name,qty) in enumerate(sp):
    ws1.append([i+1,name,qty])
    r=ws1.max_row; ws1.row_dimensions[r].height=22
    for j,cell in enumerate(ws1[r]):
        cell.font=Font(name='맑은 고딕',size=11)
        cell.border=B(T(),T(),T(),T())
        cell.alignment=Alignment(vertical='center',horizontal='left' if j==1 else 'center')
        if i%2==1: cell.fill=PatternFill('solid',fgColor='F1F8E9')
ws1.append(['',f'총 {len(sp)}종',f'총 {sum(q for _,q in sp)}개'])
r=ws1.max_row; ws1.row_dimensions[r].height=22
for cell in ws1[r]:
    cell.font=Font(name='맑은 고딕',bold=True,size=11)
    cell.fill=PatternFill('solid',fgColor='FFF176')
    cell.border=B(T('medium'),T('medium'),T('medium'),T('medium'))
    cell.alignment=Alignment(horizontal='center',vertical='center')
ws1.page_setup.orientation='portrait'; ws1.page_setup.fitToPage=True; ws1.page_setup.fitToWidth=1
wb1.save(sys.argv[2])
print(f"소분작업 저장: {sys.argv[2]} ({len(sp)}종 / 총 {sum(q for _,q in sp)}개)")

# ── 배송주소지 ────────────────────────────────────────────────────
cm = {}
for row in rows:
    addr=str(row.get('주소','') or '').strip()
    name=str(row.get('수령인명','') or row.get('수령인','') or '').strip()
    phone=str(row.get('수령인연락처','') or row.get('연락처','') or '').strip()
    raw=str(row.get('상품명(타입제거)','') or row.get('상품명','') or '').strip()
    base=int(row.get('주문수량') or row.get('수량') or 1)
    if not addr: continue
    if addr not in cm: cm[addr]={'addr':addr,'name':name,'phone':phone,'prods':[],'eco':is_eco(addr)}
    for pname,pqty in (clean_names(raw) or [(raw[:40],base)]): cm[addr]['prods'].append((pname,pqty))
eco=sorted([c for c in cm.values() if c['eco']],key=lambda x:x['addr'])
oth=sorted([c for c in cm.values() if not c['eco']],key=lambda x:x['addr'])

wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title='배송주소지'
for col,w in [('A',6),('B',50),('C',14),('D',12)]: ws2.column_dimensions[col].width=w
ws2.merge_cells('A1:D1')
c=ws2['A1']; c.value=f'배송 주소지  ─  {today_str}'
c.font=Font(name='맑은 고딕',bold=True,size=14,color='FFFFFF')
c.fill=PatternFill('solid',fgColor='1565C0')
c.alignment=Alignment(horizontal='center',vertical='center')
ws2.row_dimensions[1].height=32; ws2.append([])

def section(ws, title, color, customers):
    ws.append([title])
    r=ws.max_row; ws.merge_cells(f'A{r}:D{r}')
    c=ws.cell(r,1); c.value=title
    c.font=Font(name='맑은 고딕',bold=True,size=12,color='FFFFFF')
    c.fill=PatternFill('solid',fgColor=color)
    c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=B(T('medium'),T('medium'),T('medium'),T('medium'))
    ws.row_dimensions[r].height=26
    fc='DEEDF8' if color.startswith('1E') else 'D8EFD8'
    for num,cust in enumerate(customers,1):
        ws.append([num,cust['addr'],cust['name'],cust['phone']])
        r=ws.max_row; ws.row_dimensions[r].height=22
        for col in range(1,5):
            cell=ws.cell(r,col)
            cell.fill=PatternFill('solid',fgColor=fc)
            cell.font=Font(name='맑은 고딕',bold=(col<=2),size=10)
            cell.border=B(T('medium'),T(),T('medium') if col==1 else T(),T('medium') if col==4 else T())
            cell.alignment=Alignment(vertical='center',wrap_text=(col==2))
        for pname,pqty in cust['prods']:
            ws.append(['',f'    {pname}',f'{pqty}개',''])
            r=ws.max_row; ws.row_dimensions[r].height=20
            for col in range(1,5):
                cell=ws.cell(r,col)
                cell.font=Font(name='맑은 고딕',size=10,bold=(col==3))
                cell.border=B(T(),T(),T('medium') if col==1 else T(),T('medium') if col==4 else T())
                cell.alignment=Alignment(vertical='center',horizontal='center' if col==3 else 'left',wrap_text=(col==2))
        ws.append([])

section(ws2,f'  ★ 에코델타 지역  ({len(eco)}건)','1E88E5',eco)
ws2.append([])
section(ws2,f'  ★ 그외 지역  ({len(oth)}건)','388E3C',oth)
ws2.page_setup.orientation='portrait'; ws2.page_setup.fitToPage=True; ws2.page_setup.fitToWidth=1
wb2.save(sys.argv[3])
print(f"배송주소지 저장: {sys.argv[3]} (에코델타 {len(eco)}건 / 그외 {len(oth)}건)")
